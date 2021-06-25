import time
import datetime
import openpyxl


print("Loading EXCEL book...")
book_loading_start = time.time()
main_data_book = openpyxl.load_workbook('test case data analyst.xlsx')
print("Book loaded", time.time() - book_loading_start)
main_data_sheet = main_data_book.active
print("Sheet activated")

print("Creating new EXCEL files...")
book_for_sorting = openpyxl.Workbook()
for_sorting_sheet = book_for_sorting.active

book_for_new_data = openpyxl.Workbook()
for_data_sheet = book_for_new_data.active


def finding_first_date():
    first_date = main_data_sheet.cell(2, 4).value
    start_search_timer = time.time()
    print("Looking for the first date...")
    for row_number in range(1, main_data_sheet.max_row + 1):
        row_date = main_data_sheet.cell(row_number, 4).value
        try:
            if first_date > row_date:
                first_date = row_date
        finally:
            continue
    print("First date found in", time.time() - start_search_timer)
    return first_date


def sorting_data(first_date):
    date = first_date
    print("Sorting data...")
    start_sorting_timer = time.time()
    max_date = datetime.datetime(2020, 2, 21)
    while date < max_date:
        print(f"{date}/{max_date}")
        for row in range(1, main_data_sheet.max_row + 1):
            sheet_date = main_data_sheet.cell(row, 4).value
            # print(sheet_date)
            if sheet_date == date:
                row_list = []
                for column in range(1, 6):
                    row_list.append(main_data_sheet.cell(row, column).value)
                for_sorting_sheet.append(row_list)
        date += datetime.timedelta(days=1)
    stop_sorting_timer = time.time() - start_sorting_timer
    print("Sorting done in", stop_sorting_timer)
    print("Saving sorted result in EXCEL file by the name: 'Sorted.xlsx'")
    book_for_sorting.save('Sorted.xlsx')
    print("Saved")


def creating_cohorts():
    new_cohorts = {}
    print("Loading EXCEL book...")
    book_loading_start = time.time()
    sorted_book = openpyxl.load_workbook('Sorted.xlsx')
    sorted_sheet = sorted_book.active
    print("Laded", time.time() - book_loading_start)
    start_timer = time.time()
    for row in sorted_sheet:
        if row[2].value is True:
            try:
                new_cohorts[(row[3].value)].append(row[4].value)
            except KeyError:
                new_cohorts[(row[3].value)] = []
                new_cohorts[(row[3].value)].append(row[4].value)
    print("End of creating cohorts", time.time() - start_timer)
    sorted_book.close()
    return new_cohorts


def calculating_purchases_per_week(cohorts):
    cohorts = cohorts
    sorted_book = openpyxl.load_workbook('Sorted.xlsx')
    sorted_sheet = sorted_book.active
    purchase_amount = []
    groups_of_cohort = []
    for Date in cohorts:
        purchase_amount.append([])
        groups_of_cohort.append(Date)

    start_of_analyze = time.time()
    cohort = 0
    for Date, ID in cohorts.items():
        new_cohort_timer = time.time()
        print("New cohort analyzing...")
        week = 1
        cohort += 1
        for row_count in range(1, 120000):
            purchases = 0
            sheets_date = sorted_sheet.cell(row_count, 4).value
            try:
                sheets_date = sheets_date.date()
            except AttributeError:
                break
            cohorts_date = Date.date() + datetime.timedelta(days=7 * week)
            if cohorts_date == sheets_date:
                week += 1
                print(cohorts_date, week)
                week_timer = time.time()
                for each_id in ID:
                    inner_count = row_count
                    cohorts_day = cohorts_date.day
                    sheets_day = sorted_sheet.cell(inner_count, 4).value.date().day
                    search_timer = time.time()
                    while cohorts_day == sheets_day:
                        if each_id == sorted_sheet.cell(inner_count, 5).value:
                            purchases += 1
                            break
                        inner_count += 1
                        try:
                            sheets_day = sorted_sheet.cell(inner_count, 4).value.date().day
                        except AttributeError:
                            break
                    # print("Found 1" "in", time.time() - search_timer)
                if purchases > 0:
                    cohort_list = purchase_amount[cohort - 1]
                    print("In this week found", purchases, time.time() - week_timer)
                    cohort_list.append(purchases)
                    print(purchase_amount)
    print(purchase_amount)
    print("Data collected in", time.time() - start_of_analyze)
    return groups_of_cohort, purchase_amount


def saving_extracted_date(cohorts, purchases):
    sheet = for_data_sheet
    print("Saving cohorts...")
    row = 1
    for cohort in cohorts:
        row += 1
        sheet.cell(row, 1).value = cohort

    print("Saving purchases")
    row = 1
    for each_cohort in purchases:
        row += 1
        column = 1
        for purchases_per_week in each_cohort:
            column += 1
            sheet.cell(row, column).value = purchases_per_week

    book_for_new_data.save('Rate table.xlsx')


def analyzing_data():
    book = openpyxl.load_workbook('Rate table.xlsx')
    sheet = book.active
    weeks_values = []
    avg_week_values = []
    for row in sheet:
        index = 0
        for cell in row:
            if type(cell.value) is int:
                try:
                    weeks_values[index].append(cell.value)
                except IndexError:
                    weeks_values.append([])
                    weeks_values[index].append(cell.value)
                index += 1
    for week in weeks_values:
        dev = 0
        summ = 0
        for w in week:
            dev += 1
            summ += w
        avg_week_values.append(int(summ/dev))

    avg_cr_period = 0
    for num in range(len(avg_week_values)):
        cr = (avg_week_values[0] - avg_week_values[num]) / avg_week_values[0]
        if cr > 0.50:
            avg_cr_period = num + 1
            break

    cltv_avg = 5 * avg_cr_period

    month_cr = ((avg_week_values[0] - avg_week_values[3]) / avg_week_values[0])

    cltv_month = int(4.99 / month_cr)

    month_purch = avg_week_values[0] + avg_week_values[1] + avg_week_values[2] + avg_week_values[3]
    romi = int((month_purch * 4.99 - avg_week_values[0] * 6) / avg_week_values[0] * 6)

    w_b = openpyxl.Workbook()
    w_s = w_b.active

    w_s.append(["CLV in avg week", cltv_avg])
    w_s.append(["CLV in one month", cltv_month])
    w_s.append(["ROMI", romi])
    x = ["AVG week purchases:"]
    for i in avg_week_values:
        x.append(i)
    w_s.append(x)
    w_s.append([])
    w_s.append(["Cohorts", "1-st week", "2-nd week", "3-d week", "4-th week", "5-th week", "6-th week", "7-th week"])
    for row in sheet:
        list_row = []
        if row[2].value is not None:
            for cell in row:
                list_row.append(cell.value)
            w_s.append(list_row)

    w_b.save("Finish.xlsx")
    w_b.close()


def run():
    first_date = finding_first_date()
    print(first_date)
    sorting_data(first_date)
    cohorts = creating_cohorts()
    list_of_cohorts, list_of_purchases = calculating_purchases_per_week(cohorts)
    saving_extracted_date(list_of_cohorts, list_of_purchases)
    analyzing_data()


if __name__ == '__main__':
    run()


main_data_book.close()
book_for_sorting.close()
book_for_new_data.close()
